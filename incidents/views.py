import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField, Q
from .models import Ticket, TicketHistorique
from .forms import TicketForm, ReaffectationForm
from accounts.models import Utilisateur
from core.models import Zone


@login_required
def ticket_list(request):
    user = request.user
    if user.role == 'helpdesk':
        tickets = Ticket.objects.all().order_by('-date_creation')
    elif user.role in ['technicien', 'ingenieur']:
        tickets = Ticket.objects.filter(assigne_a=user).order_by('-date_creation')
    elif user.role == 'superviseur':
        tickets = Ticket.objects.all().order_by('-date_creation')
    else:
        tickets = Ticket.objects.none()

    statut = request.GET.get('statut', '')
    priorite = request.GET.get('priorite', '')
    zone_id = request.GET.get('zone', '')
    q = request.GET.get('q', '').strip()

    if statut:
        tickets = tickets.filter(statut=statut)
    if priorite:
        tickets = tickets.filter(priorite=priorite)
    if zone_id:
        tickets = tickets.filter(zone_id=zone_id)
    if q:
        tickets = tickets.filter(
            Q(numero_vol__icontains=q) | Q(compagnie_aerienne__icontains=q)
        )

    context = {
        'tickets': tickets,
        'zones': Zone.objects.all().order_by('nom'),
        'statut_choices': Ticket.STATUT_CHOICES,
        'priorite_choices': Ticket.PRIORITE_CHOICES,
        'filtre_statut': statut,
        'filtre_priorite': priorite,
        'filtre_zone': zone_id,
        'filtre_q': q,
    }
    return render(request, 'incidents/ticket_list.html', context)


@login_required
def ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    historique = ticket.historique.all() if request.user.role == 'superviseur' else None
    return render(request, 'incidents/ticket_detail.html', {
        'ticket': ticket,
        'historique': historique,
    })


def affecter_agent(ticket):
    """Trouve un agent disponible dans la zone du ticket, selon la priorite."""
    if ticket.priorite in ['P1', 'P2']:
        role_cible = 'ingenieur'
    else:
        role_cible = 'technicien'

    agent = Utilisateur.objects.filter(
        role=role_cible,
        zone=ticket.zone,
        disponibilite='disponible'
    ).first()

    if agent:
        agent.disponibilite = 'occupe'
        agent.save()

    return agent


@login_required
def ticket_create(request):
    if request.user.role != 'helpdesk':
        messages.error(request, "Seul un agent Helpdesk peut creer un ticket.")
        return redirect('incidents:ticket_list')

    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.cree_par = request.user

            agent = affecter_agent(ticket)
            if agent:
                ticket.assigne_a = agent
                ticket.statut = 'affecte'
            else:
                ticket.statut = 'ouvert'

            ticket.save()

            TicketHistorique.objects.create(
                ticket=ticket,
                utilisateur=request.user,
                action='Creation du ticket',
                ancien_statut='',
                nouveau_statut=ticket.statut,
                commentaire=f"Affecte a {agent}" if agent else "Aucun agent disponible dans la zone"
            )

            if agent:
                messages.success(request, f"Ticket cree et affecte a {agent}.")
            else:
                messages.warning(request, "Ticket cree mais aucun agent disponible dans cette zone.")

            return redirect('incidents:ticket_detail', pk=ticket.pk)
    else:
        form = TicketForm()

    return render(request, 'incidents/ticket_form.html', {'form': form})


@login_required
def ticket_resoudre(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if ticket.assigne_a != request.user:
        messages.error(request, "Ce ticket ne vous est pas assigne.")
        return redirect('incidents:ticket_list')

    ancien_statut = ticket.statut
    ticket.statut = 'resolu'
    ticket.save()

    request.user.disponibilite = 'disponible'
    request.user.save()

    TicketHistorique.objects.create(
        ticket=ticket,
        utilisateur=request.user,
        action='Resolution',
        ancien_statut=ancien_statut,
        nouveau_statut='resolu',
        commentaire='Marque comme resolu par ' + str(request.user)
    )
    messages.success(request, "Ticket marque comme resolu.")
    return redirect('incidents:ticket_detail', pk=ticket.pk)


@login_required
def ticket_escalader(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if ticket.assigne_a != request.user or request.user.role != 'technicien':
        messages.error(request, "Action non autorisee.")
        return redirect('incidents:ticket_list')

    ancien_statut = ticket.statut

    request.user.disponibilite = 'disponible'
    request.user.save()

    ticket.statut = 'escalade'
    ticket.assigne_a = None
    ticket.save()

    TicketHistorique.objects.create(
        ticket=ticket,
        utilisateur=request.user,
        action='Escalade',
        ancien_statut=ancien_statut,
        nouveau_statut='escalade',
        commentaire=f"Signale par {request.user} - en attente de reaffectation par le Helpdesk"
    )
    messages.warning(request, "Ticket escalade. En attente de reaffectation par le Helpdesk.")
    return redirect('incidents:ticket_detail', pk=ticket.pk)


@login_required
def ticket_reaffecter(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role != 'helpdesk':
        messages.error(request, "Seul le Helpdesk peut reaffecter un ticket.")
        return redirect('incidents:ticket_list')

    if ticket.statut != 'escalade':
        messages.error(request, "Ce ticket n'est pas en attente de reaffectation.")
        return redirect('incidents:ticket_detail', pk=ticket.pk)

    if request.method == 'POST':
        form = ReaffectationForm(request.POST, zone=ticket.zone)
        if form.is_valid():
            ingenieur = form.cleaned_data['ingenieur']
            ancien_statut = ticket.statut

            ticket.assigne_a = ingenieur
            ticket.statut = 'affecte'
            ticket.save()

            ingenieur.disponibilite = 'occupe'
            ingenieur.save()

            TicketHistorique.objects.create(
                ticket=ticket,
                utilisateur=request.user,
                action='Reaffectation',
                ancien_statut=ancien_statut,
                nouveau_statut='affecte',
                commentaire=f"Reaffecte par {request.user} vers {ingenieur}"
            )
            messages.success(request, f"Ticket reaffecte a {ingenieur}.")
            return redirect('incidents:ticket_detail', pk=ticket.pk)
    else:
        form = ReaffectationForm(zone=ticket.zone)

    return render(request, 'incidents/ticket_reaffecter.html', {'form': form, 'ticket': ticket})


@login_required
def ticket_cloturer(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role != 'helpdesk':
        messages.error(request, "Seul le Helpdesk peut cloturer un ticket.")
        return redirect('incidents:ticket_list')

    if request.method == 'POST':
        commentaire = request.POST.get('commentaire', '')
        ancien_statut = ticket.statut
        ticket.statut = 'cloture'
        ticket.commentaire_cloture = commentaire
        ticket.date_cloture = timezone.now()
        ticket.save()

        TicketHistorique.objects.create(
            ticket=ticket,
            utilisateur=request.user,
            action='Cloture',
            ancien_statut=ancien_statut,
            nouveau_statut='cloture',
            commentaire=commentaire
        )
        messages.success(request, "Ticket cloture.")
        return redirect('incidents:ticket_detail', pk=ticket.pk)

    return render(request, 'incidents/ticket_cloture_form.html', {'ticket': ticket})


def _get_stats(tickets):
    """Calcule les stats partagees entre le dashboard web et les exports."""
    total = tickets.count()
    par_statut = list(tickets.values('statut').annotate(total=Count('id')))
    par_zone = list(tickets.values('zone__nom').annotate(total=Count('id')))
    par_priorite = list(tickets.values('priorite').annotate(total=Count('id')))

    tickets_resolus = tickets.filter(statut='cloture', date_cloture__isnull=False)
    duree_moyenne = tickets_resolus.annotate(
        duree=ExpressionWrapper(F('date_cloture') - F('date_creation'), output_field=DurationField())
    ).aggregate(moyenne=Avg('duree'))['moyenne']

    return total, par_statut, par_zone, par_priorite, duree_moyenne


def _format_duree(td):
    """Formate un timedelta en texte lisible : '1j 3h 12min' ou '45min' etc."""
    if td is None:
        return None
    total_seconds = int(td.total_seconds())
    jours, reste = divmod(total_seconds, 86400)
    heures, reste = divmod(reste, 3600)
    minutes, _ = divmod(reste, 60)

    parts = []
    if jours:
        parts.append(f"{jours}j")
    if heures:
        parts.append(f"{heures}h")
    if minutes or not parts:
        parts.append(f"{minutes}min")
    return " ".join(parts)


@login_required
def dashboard_superviseur(request):
    if request.user.role != 'superviseur':
        messages.error(request, "Acces reserve au superviseur.")
        return redirect('incidents:ticket_list')

    tickets = Ticket.objects.all()
    total, par_statut, par_zone, par_priorite, duree_moyenne = _get_stats(tickets)

    chart_statut = {
        'labels': [s['statut'] for s in par_statut],
        'data': [s['total'] for s in par_statut],
    }
    chart_zone = {
        'labels': [z['zone__nom'] or '-' for z in par_zone],
        'data': [z['total'] for z in par_zone],
    }
    chart_priorite = {
        'labels': [p['priorite'] for p in par_priorite],
        'data': [p['total'] for p in par_priorite],
    }

    return render(request, 'incidents/dashboard.html', {
        'total': total,
        'par_statut': par_statut,
        'par_zone': par_zone,
        'par_priorite': par_priorite,
        'duree_moyenne': _format_duree(duree_moyenne),
        'tickets_recents': tickets.order_by('-date_creation')[:10],
        'chart_statut_json': json.dumps(chart_statut),
        'chart_zone_json': json.dumps(chart_zone),
        'chart_priorite_json': json.dumps(chart_priorite),
    })


@login_required
def dashboard_export_excel(request):
    if request.user.role != 'superviseur':
        messages.error(request, "Acces reserve au superviseur.")
        return redirect('incidents:ticket_list')

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    tickets = Ticket.objects.all()
    total, par_statut, par_zone, par_priorite, duree_moyenne = _get_stats(tickets)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport ITmanager"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws.append(["Rapport superviseur - ITmanager"])
    ws.append([f"Genere le {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.append([])
    ws.append([f"Total tickets : {total}"])
    if duree_moyenne:
        ws.append([f"Duree moyenne de resolution : {_format_duree(duree_moyenne)}"])
    ws.append([])

    def add_section(title, rows, headers):
        ws.append([title])
        ws.append(headers)
        for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
        for r in rows:
            ws.append(r)
        ws.append([])

    add_section("Par statut", [[s['statut'], s['total']] for s in par_statut], ["Statut", "Total"])
    add_section("Par zone", [[z['zone__nom'] or '-', z['total']] for z in par_zone], ["Zone", "Total"])
    add_section("Par priorite", [[p['priorite'], p['total']] for p in par_priorite], ["Priorite", "Total"])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_itmanager_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


def _make_chart_image(labels, values, title, chart_type='bar'):
    """Genere un graphique matplotlib et le retourne en buffer PNG."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import io

    colors_palette = ['#2563EB', '#60A5FA', '#93C5FD', '#DC2626', '#F59E0B', '#16A34A']

    fig, ax = plt.subplots(figsize=(6, 3))
    if chart_type == 'bar':
        ax.bar(labels, values, color=colors_palette[:len(labels)])
        ax.set_ylabel('Nombre de tickets')
        for i, v in enumerate(values):
            ax.text(i, v + max(values, default=1) * 0.02, str(v), ha='center', fontsize=9)
    else:
        ax.pie(values, labels=labels, autopct='%1.0f%%', colors=colors_palette[:len(labels)])

    ax.set_title(title, fontsize=12, fontweight='bold')
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


@login_required
def dashboard_export_pdf(request):
    if request.user.role != 'superviseur':
        messages.error(request, "Acces reserve au superviseur.")
        return redirect('incidents:ticket_list')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    tickets = Ticket.objects.all()
    total, par_statut, par_zone, par_priorite, duree_moyenne = _get_stats(tickets)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Rapport superviseur - ITmanager", styles['Title']))
    elements.append(Paragraph(f"Genere le {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"<b>Total tickets :</b> {total}", styles['Normal']))
    if duree_moyenne:
        elements.append(Paragraph(f"<b>Duree moyenne de resolution :</b> {_format_duree(duree_moyenne)}", styles['Normal']))
    elements.append(Spacer(1, 0.7*cm))

    def make_table(title, rows, headers):
        elements.append(Paragraph(title, styles['Heading2']))
        data = [headers] + rows
        t = Table(data, colWidths=[9*cm, 4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E5EA')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F8FA')]),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))

    def make_chart(labels, values, title, chart_type='bar'):
        if not labels:
            return
        img_buf = _make_chart_image(labels, values, title, chart_type)
        elements.append(Image(img_buf, width=14*cm, height=7*cm))
        elements.append(Spacer(1, 0.6*cm))

    statut_labels = [s['statut'] for s in par_statut]
    statut_values = [s['total'] for s in par_statut]
    zone_labels = [z['zone__nom'] or '-' for z in par_zone]
    zone_values = [z['total'] for z in par_zone]
    priorite_labels = [p['priorite'] for p in par_priorite]
    priorite_values = [p['total'] for p in par_priorite]

    make_table("Par statut", [[s, str(v)] for s, v in zip(statut_labels, statut_values)], ["Statut", "Total"])
    make_chart(statut_labels, statut_values, "Repartition par statut", 'pie')

    make_table("Par zone", [[z, str(v)] for z, v in zip(zone_labels, zone_values)], ["Zone", "Total"])
    make_chart(zone_labels, zone_values, "Repartition par zone", 'bar')

    make_table("Par priorite", [[p, str(v)] for p, v in zip(priorite_labels, priorite_values)], ["Priorite", "Total"])
    make_chart(priorite_labels, priorite_values, "Repartition par priorite", 'bar')

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapport_itmanager_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response